import os
import glob
import time
import openpyxl
import subprocess
import inspect
from datetime import datetime


# 获取项目路径
current_path = inspect.getfile(inspect.currentframe())
projectDir = os.path.abspath(os.path.join(os.path.dirname(current_path), "../"))

# 设置文件夹路径和求解器路径
instances_path = "../Data/benchmark/SAT"  # 算例文件夹路径
solver_path = "build/cirsat"  # 求解器路径

# 设置输出的 Excel 文件路径到上一级目录的 result 文件夹
result_dir = os.path.join(os.path.dirname(os.getcwd()), 'result')  # 上一级目录的 result 文件夹
os.makedirs(result_dir, exist_ok=True)  # 创建 result 文件夹（如果不存在）

result_file = os.path.join(result_dir, 'timing_results.xlsx')  # 输出文件路径

# 设置超时时间为3600秒（1小时）
timeout_duration = 3600  

# 检查 Excel 文件是否存在，如果存在则打开，否则创建新的 Excel 文件
if os.path.exists(result_file):
    wb = openpyxl.load_workbook(result_file)
else:
    wb = openpyxl.Workbook()

# 获取 .aag 文件
class DATA:
    def __init__(self):
        self.dataPath = ""

    def getFile(self, path, condition):
        g = os.walk(path)

        pathfile = []
        for path, dir_list, file_list in g:
            for condi in condition:
                for file_name in glob.glob(os.path.join(path, condi)):
                    pathfile.append(file_name)
        return pathfile

    def getFileName(self, file):
        return file.split("/")[-1]


conditions = ["*.aag"]
files = DATA().getFile(instances_path, conditions)

# 遍历每个文件夹
folder_results = {}  # 用来存储文件夹的结果，以文件夹名为键

for file in files:
    file_name = DATA().getFileName(file).split(".aag")[0]
    folder_name = os.path.basename(os.path.dirname(file))  # 获取当前文件夹的名字

    # 为每个文件夹创建一个新的工作表，或者在现有的表格中添加数据
    if folder_name not in folder_results:
        # 如果该文件夹还没有工作表，则创建一个
        if folder_name not in wb.sheetnames:
            sheet = wb.create_sheet(title=folder_name)
            sheet.append(["算例名", "运行时间(毫秒)", "结果ANS", "CHECK"])  # 设置表头
        folder_results[folder_name] = wb[folder_name]
    else:
        # 如果该文件夹已经有工作表，获取该工作表
        sheet = folder_results[folder_name]

    # 获取文件的路径
    aag_file_path = file

    # 记录开始时间（精确到毫秒）
    start_time_ms = int(time.time() * 1000)

    # 创建临时文件存储求解器的输出
    temp_output = os.path.join(os.path.dirname(aag_file_path), 'temp_output.txt')

    try:
        # 执行求解器并限制超时，超过时间则停止并输出timeout
        result = subprocess.run(
            [solver_path, "-aiger", aag_file_path],
            timeout=timeout_duration,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE
        )
        with open(temp_output, 'wb') as f:
            f.write(result.stdout)
            f.write(result.stderr)
    except subprocess.TimeoutExpired:
        # 如果超时，则记录超时
        sheet.append([file_name, 'Timeout', 'Timeout'])
        continue

    # 记录结束时间（精确到毫秒）
    end_time_ms = int(time.time() * 1000)

    # 计算运行时间（毫秒）
    run_time_ms = end_time_ms - start_time_ms

    # 从求解器输出中提取ANS字段
    with open(temp_output, 'r') as f:
        output = f.read()
        # 提取ANS: UNSAT 或 ANS: SAT
        ans_status = "ANS: UNSAT" if "ANS: UNSAT" in output else "ANS: SAT" if "ANS: SAT" in output else "Unknown"

        # 提取 check 字段
        check_status = "check: UNSAT" if "check: UNSAT" in output else "check: SAT" if "check: SAT" in output else "Unknown"

    # 将结果写入工作表
    sheet.append([file_name, run_time_ms, ans_status, check_status])

    # 清理临时文件
    os.remove(temp_output)

# 保存 Excel 文件
wb.save(result_file)
print(f"Excel 文件已保存到: {result_file}")
